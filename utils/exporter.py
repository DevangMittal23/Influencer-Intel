import io
import pandas as pd
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

VERDICT_FILLS = {
    "TRUE":           PatternFill("solid", fgColor="0D2818"),
    "FALSE":          PatternFill("solid", fgColor="2D0F1A"),
    "MISLEADING":     PatternFill("solid", fgColor="2D1F00"),
    "PARTIALLY_TRUE": PatternFill("solid", fgColor="1A2020"),
    "UNVERIFIABLE":   PatternFill("solid", fgColor="1A1D27"),
    "SKIPPED":        PatternFill("solid", fgColor="1A1D27"),
}
VERDICT_FONTS = {
    "TRUE":           Font(color="00D4AA", bold=True),
    "FALSE":          Font(color="FF4B6E", bold=True),
    "MISLEADING":     Font(color="FFB800", bold=True),
    "PARTIALLY_TRUE": Font(color="40C4AA", bold=True),
    "UNVERIFIABLE":   Font(color="6B7280", bold=True),
    "SKIPPED":        Font(color="6B7280"),
}


def _count_verdicts(fact_checks: list) -> dict:
    counts = {"TRUE": 0, "FALSE": 0, "MISLEADING": 0, "UNVERIFIABLE": 0, "PARTIALLY_TRUE": 0}
    for fc in fact_checks:
        v = fc.get("verdict", "")
        if v in counts:
            counts[v] += 1
    return counts


def export_to_csv(results: list) -> bytes:
    rows = []
    for r in results:
        fact_checks = r.get("fact_checks", [])
        counts = _count_verdicts(fact_checks)
        entities_str = "; ".join(
            f"{e.get('name')} ({e.get('stance')})" for e in r.get("entities", [])
        )
        cf = r.get("campaign_fit", {})
        risk = r.get("risk_assessment", {})
        base = {
            "Source": r.get("source_name", ""),
            "Core Narrative": r.get("core_narrative", ""),
            "Intent": r.get("intent", ""),
            "Sentiment": r.get("sentiment_overall", ""),
            "Fit Score": cf.get("score", ""),
            "Fit Label": cf.get("label", ""),
            "Risk Level": risk.get("risk_level", "N/A"),
            "Risk Score": risk.get("risk_score", 0),
            "Needs Review": "Yes" if risk.get("needs_human_review", False) else "No",
            "Entities": entities_str,
            "True Claims": counts["TRUE"],
            "False Claims": counts["FALSE"],
            "Misleading": counts["MISLEADING"],
            "Unverifiable": counts["UNVERIFIABLE"],
            "Provider": r.get("provider_used", ""),
        }
        if fact_checks:
            for fc in fact_checks:
                rows.append({**base,
                    "Claim": fc.get("claim", ""),
                    "Claim Type": fc.get("claim_type", ""),
                    "Verdict": fc.get("verdict", ""),
                    "Confidence": fc.get("confidence", ""),
                    "Reasoning": fc.get("reasoning", ""),
                    "Correct Fact": fc.get("correct_fact", "") or "",
                    "Source URL": fc.get("source_url", "") or "",
                })
        else:
            rows.append({**base,
                "Claim": "", "Claim Type": "", "Verdict": "",
                "Confidence": "", "Reasoning": "", "Correct Fact": "", "Source URL": "",
            })

    buf = io.BytesIO()
    pd.DataFrame(rows).to_csv(buf, index=False)
    return buf.getvalue()


def export_to_excel(results: list) -> bytes:
    summary_rows, fc_rows, entity_rows = [], [], []

    for r in results:
        fact_checks = r.get("fact_checks", [])
        counts = _count_verdicts(fact_checks)
        cf = r.get("campaign_fit", {})
        risk = r.get("risk_assessment", {})
        alignment = r.get("campaign_alignment", {})
        warnings = "; ".join(w for w in r.get("content_warnings", []) if w and w != "none")

        summary_rows.append({
            "Source": r.get("source_name", ""),
            "Core Narrative": r.get("core_narrative", ""),
            "Intent": r.get("intent", ""),
            "Campaign Fit Score": cf.get("score", ""),
            "Campaign Fit Label": cf.get("label", ""),
            "Overall Sentiment": r.get("sentiment_overall", ""),
            "Content Warnings": warnings,
            "Total Claims": len(fact_checks),
            "True Claims": counts["TRUE"],
            "False Claims": counts["FALSE"],
            "Misleading": counts["MISLEADING"],
            "Unverifiable": counts["UNVERIFIABLE"],
            "Risk Level": risk.get("risk_level", "N/A"),
            "Risk Score": risk.get("risk_score", 0),
            "Needs Review": "Yes" if risk.get("needs_human_review", False) else "No",
            "Campaign Aligned": "Yes" if alignment.get("aligned", False) else "No",
            "Alignment Summary": alignment.get("alignment_summary", ""),
            "Red Flags": " | ".join(alignment.get("red_flags", [])),
            "Provider Used": r.get("provider_used", ""),
        })

        for fc in fact_checks:
            fc_rows.append({
                "Source": r.get("source_name", ""),
                "Claim": fc.get("claim", ""),
                "Type": fc.get("claim_type", ""),
                "Verdict": fc.get("verdict", ""),
                "Confidence": fc.get("confidence", ""),
                "Reasoning": fc.get("reasoning", ""),
                "Correct Fact": fc.get("correct_fact", "") or "",
                "Source URL": fc.get("source_url", "") or "",
            })

        for e in r.get("entities", []):
            entity_rows.append({
                "Source": r.get("source_name", ""),
                "Entity Name": e.get("name", ""),
                "Type": e.get("type", ""),
                "Stance": e.get("stance", ""),
                "Context": e.get("context", ""),
            })

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df_sum = pd.DataFrame(summary_rows)
        df_fc = pd.DataFrame(fc_rows) if fc_rows else pd.DataFrame(
            columns=["Source", "Claim", "Type", "Verdict", "Confidence", "Reasoning", "Correct Fact", "Source URL"]
        )
        df_ent = pd.DataFrame(entity_rows) if entity_rows else pd.DataFrame(
            columns=["Source", "Entity Name", "Type", "Stance", "Context"]
        )

        df_sum.to_excel(writer, sheet_name="Summary", index=False)
        df_fc.to_excel(writer, sheet_name="Fact Checks", index=False)
        df_ent.to_excel(writer, sheet_name="Entities", index=False)

        wb = writer.book
        ws_sum = wb["Summary"]

        # Color: Campaign Fit Score
        for cell in ws_sum[1]:
            if cell.value == "Campaign Fit Score":
                for row in ws_sum.iter_rows(min_row=2, min_col=cell.column, max_col=cell.column):
                    for c in row:
                        if isinstance(c.value, (int, float)):
                            c.font = Font(
                                color="00D4AA" if c.value >= 80 else "FFB800" if c.value >= 50 else "FF4B6E",
                                bold=True
                            )
                break

        # Color: Risk Level column
        for cell in ws_sum[1]:
            if cell.value == "Risk Level":
                risk_fills = {
                    "HIGH":   PatternFill("solid", fgColor="FF4B6E"),
                    "MEDIUM": PatternFill("solid", fgColor="FFB800"),
                    "LOW":    PatternFill("solid", fgColor="00D4AA"),
                }
                for row in ws_sum.iter_rows(min_row=2, min_col=cell.column, max_col=cell.column):
                    for c in row:
                        if str(c.value) in risk_fills:
                            c.fill = risk_fills[str(c.value)]
                break

        # Color: Verdict column in Fact Checks
        ws_fc = wb["Fact Checks"]
        for cell in ws_fc[1]:
            if cell.value == "Verdict":
                for row in ws_fc.iter_rows(min_row=2, min_col=cell.column, max_col=cell.column):
                    for c in row:
                        v = str(c.value or "")
                        if v in VERDICT_FILLS:
                            c.fill = VERDICT_FILLS[v]
                            c.font = VERDICT_FONTS.get(v, Font())
                break

        # Sheet 4: Review Queue
        ws_review = wb.create_sheet("Review Queue")
        review_headers = [
            "Source", "Risk Level", "Risk Score",
            "Review Reason", "Risk Factors",
            "False Claims Count", "Campaign Fit Score",
        ]
        ws_review.append(review_headers)
        for r in results:
            risk = r.get("risk_assessment", {})
            if risk.get("needs_human_review", False):
                false_count = sum(
                    1 for fc in r.get("fact_checks", [])
                    if fc.get("verdict") in ["FALSE", "MISLEADING"]
                )
                ws_review.append([
                    r.get("source_name", ""),
                    risk.get("risk_level", ""),
                    risk.get("risk_score", 0),
                    risk.get("review_reason", ""),
                    " | ".join(risk.get("risk_factors", [])),
                    false_count,
                    r.get("campaign_fit", {}).get("score", 0),
                ])

        # Auto-width all sheets
        for ws in wb.worksheets:
            for col in ws.columns:
                max_len = max((len(str(c.value or "")) for c in col), default=10)
                ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 60)

    return buf.getvalue()
